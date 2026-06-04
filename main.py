"""
Live Chat Backend
─────────────────────────
Run:  uvicorn main:app --reload

Endpoints:
  POST   /chat                  – save a message
  GET    /conversations         – list all sessions (JSON)
  GET    /conversations/{id}    – single session messages (JSON)
  GET    /export/excel          – download all conversations as .xlsx
  DELETE /conversations/{id}    – delete a session

Storage: conversations.json  (flat file, no DB needed)
"""

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from typing import Optional, List
import json, os, io, re
from datetime import datetime, timezone

# ─── openpyxl for Excel export ───────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─── Groq AI ──────────────────────────────────────────────────────────────────
try:
    from groq import Groq
    _groq_key = os.environ.get("GROQ_API_KEY", "")
    groq_client = Groq(api_key=_groq_key) if _groq_key else None
except ImportError:
    groq_client = None

SYSTEM_PROMPT = """You are a helpful, friendly customer support AI assistant.
- Keep answers concise (2-4 sentences) unless the user needs more detail
- Be warm, professional, and empathetic
- Respond in the same language the user writes in
- If you don't know specific business details (prices, stock, order info), say so honestly and suggest contacting a human agent
- Never make up information you are not certain about"""

# ─── App ─────────────────────────────────────────────────────────────────────
app = FastAPI(title="Live Chat API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # restrict to your domain in production
    allow_methods=["*"],
    allow_headers=["*"],
)

# ─── Storage file ─────────────────────────────────────────────────────────────
DATA_FILE = "conversations.json"

def _load() -> dict:
    if not os.path.exists(DATA_FILE):
        return {}
    with open(DATA_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def _save(data: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

# ─── Models ───────────────────────────────────────────────────────────────────
class ChatMessage(BaseModel):
    session_id: str
    role: str           # "user" or "bot"
    message: str
    page: Optional[str] = None

class HistoryItem(BaseModel):
    role: str           # "user" or "assistant"
    content: str

class AiChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None
    history: Optional[List[HistoryItem]] = None

# ─── Routes ───────────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "service": "Live Chat API"}


@app.get("/widget.js")
def serve_widget():
    """Serve the embeddable chat widget plugin."""
    return FileResponse(
        "widget.js",
        media_type="application/javascript",
        headers={"Cache-Control": "public, max-age=3600"},
    )


@app.post("/ai-chat")
async def ai_chat(req: AiChatRequest):
    """Call Groq AI and return a reply. Requires GROQ_API_KEY env var."""
    if not groq_client:
        raise HTTPException(
            status_code=503,
            detail="AI service not configured. Set GROQ_API_KEY environment variable."
        )
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]
    if req.history:
        for h in req.history[-10:]:          # keep last 10 turns for context
            messages.append({"role": h.role, "content": h.content})
    messages.append({"role": "user", "content": req.message})

    completion = groq_client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=messages,
        max_tokens=400,
        temperature=0.7,
    )
    reply = completion.choices[0].message.content.strip()
    return {"reply": reply}


@app.post("/chat", status_code=201)
def save_message(msg: ChatMessage):
    """Receive a single chat message and append to the session."""
    data = _load()

    if msg.session_id not in data:
        data[msg.session_id] = {
            "session_id": msg.session_id,
            "started_at": datetime.now(timezone.utc).isoformat(),
            "page": msg.page or "",
            "messages": []
        }

    entry = {
        "role": msg.role,
        "message": msg.message,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }
    data[msg.session_id]["messages"].append(entry)
    data[msg.session_id]["last_active"] = entry["timestamp"]

    _save(data)
    return {"saved": True, "session_id": msg.session_id}


@app.get("/conversations")
def list_conversations():
    """Return a summary of all sessions."""
    data = _load()
    summaries = []
    for sid, sess in data.items():
        msgs = sess.get("messages", [])
        user_msgs = [m for m in msgs if m["role"] == "user"]
        summaries.append({
            "session_id": sid,
            "started_at": sess.get("started_at"),
            "last_active": sess.get("last_active"),
            "page": sess.get("page", ""),
            "total_messages": len(msgs),
            "user_messages": len(user_msgs),
            "first_user_message": user_msgs[0]["message"] if user_msgs else ""
        })
    # newest first
    summaries.sort(key=lambda x: x.get("last_active") or "", reverse=True)
    return {"count": len(summaries), "sessions": summaries}


@app.get("/conversations/{session_id}")
def get_conversation(session_id: str):
    """Return full message history for one session."""
    data = _load()
    if session_id not in data:
        raise HTTPException(status_code=404, detail="Session not found")
    return data[session_id]


@app.delete("/conversations/{session_id}")
def delete_conversation(session_id: str):
    data = _load()
    if session_id not in data:
        raise HTTPException(status_code=404, detail="Session not found")
    del data[session_id]
    _save(data)
    return {"deleted": True}


@app.get("/export/excel")
def export_excel():
    """
    Export ALL conversations to an .xlsx file.
    Each row = one message.
    Columns: Session ID | Started At | Page | Role | Message | Timestamp
    """
    if not EXCEL_OK:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed. Run: pip install openpyxl"
        )

    data = _load()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conversations"

    # ── Header styling ──
    green  = "1e8a44"
    header_fill = PatternFill("solid", fgColor=green)
    white_bold  = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    headers = ["Session ID", "Started At", "Page / URL", "Role", "Message", "Message Timestamp"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font  = white_bold
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Data rows ──
    row_num = 2
    user_fill = PatternFill("solid", fgColor="E8F5EC")
    bot_fill  = PatternFill("solid", fgColor="F5F5F5")

    for sid, sess in data.items():
        started   = sess.get("started_at", "")
        page      = sess.get("page", "")
        messages  = sess.get("messages", [])

        for m in messages:
            fill = user_fill if m["role"] == "user" else bot_fill
            values = [sid, started, page, m["role"], m["message"], m.get("timestamp", "")]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill   = fill
                cell.border = thin_border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_num += 1

    # ── Column widths ──
    col_widths = [26, 22, 30, 8, 60, 22]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"   # freeze header row

    # ── Stream file ──
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"livechat_chats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
